"""
BSO → Next Project Transformer
================================
Läser en Excel-fil exporterad av BSOExcel_export_till_next_2.0 och
producerar färdiga Nuvo-importfiler för Next Project.

Användning:
    python bso_to_next_transformer.py <bso_export.xlsx> [--output-dir <katalog>]

Krav:
    pip install openpyxl pandas

Output (numrerade för att köras i rätt importordning):
    01_Timpriser.xlsx       → Resurser / Nuvo: next_artikel (kategori Tidspris)
    02_Prislista.xlsx       → Prislista / Nuvo: next_artikel
    03_Projekt.xlsx         → Projekt / Nuvo: next_projekt
    04_ÄTA.xlsx             → Avvikelser → ÄTA-headers / Nuvo: next_additionalworkorder
    05_Kostnader.xlsx       → Bokförda kostnader / Nuvo: next_kostnad (import 13)
    06_Timmar.xlsx          → Bokförda timmar / Nuvo: next_kostnad (konto 7010)
    TRANSFORMATIONSLOGG.xlsx → Varningar, okände värden, rader som hoppats över

Kolumnkarta (BSO → Next Nuvo):
    Se BSO_COLUMN_MAPS nedan. Baserat på SQL-queries ur BSOExcel.exe (reverse-engineered).

Version: 1.0  /  2026-05-28
"""

import sys
import os
import argparse
import logging
from datetime import datetime, date
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Saknade paket. Kör: pip install pandas openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURATION — Justera dessa mappningar per kund vid behov
# ─────────────────────────────────────────────────────────────────────────────

# BSO cDebForm (Ersättningsform) → Next giltigt värde
ERSATTNINGSFORM_MAP = {
    "F":              "Fastpris",
    "FAST":           "Fastpris",
    "FASTPRIS":       "Fastpris",
    "L":              "Löpande räkning",
    "LOP":            "Löpande räkning",
    "LOPANDE":        "Löpande räkning",
    "LÖPANDE":        "Löpande räkning",
    "LR":             "Löpande räkning",
    "A":              "à-pris",
    "APRIS":          "à-pris",
    "À-PRIS":         "à-pris",
    "Ö":              "Övrigt",
    "OVRIGT":         "Övrigt",
    "ÖVRIGT":         "Övrigt",
    # Lägg till kundspecifika värden här
}

# BSO cStatus → Next Projektstatus (anpassa per kunds Next-installation)
# Dessa är Next-standardvärden — kontrollera mot GET /status/project/ per kund
BSO_STATUS_MAP = {
    "10": "Anbud / återkoppling",
    "20": "Nytt",
    "30": "Beställt",
    "40": "Pågående",
    "50": "Pågående",
    "60": "Utfört",
    "70": "Avslutat",
    "80": "Slutfakturerat",
    "90": "Avslutat",
    # BSO-exporten sätter hårdkodat '50' som Status i projektexporten.
    # BSStatus = originalt BSO-värde, behålls som notering.
}

# BSO Resurstyp → Next Artikelkategori
RESURSTYP_MAP = {
    "MASKIN":   "Egna maskiner",
    "FORDON":   "Fordon",
    "MATERIAL": "Eget material",
    "REDSKAP":  "Verktyg",
    "VERKTYG":  "Verktyg",
    "UE":       "Eget material",
    None:       "Eget material",
    "":         "Eget material",
}

# BSO ÄTA cDocType → Next ÄTA-Typ
AVVTYP_MAP = {
    "AVVE": "ÄTA",           # Extern ÄTA
    "AVVI": "Intern ÄTA",    # Intern avvikelse
    "FS":   "Regiberedning", # Fakturaspecifikation / löpande
}

# Fallback-status för Next Projektstatus när BSO-värdet är okänt
DEFAULT_PROJECT_STATUS = "Pågående"

# Konton att hoppa över vid kostnadsimport (balansräkningskonton)
SKIP_ACCOUNTS_PREFIX = ("1", "2", "8", "9")

# Minsta projektstart-datum — projekt äldre än detta klassas som historiska
CUTOFF_DATE_DEFAULT = "2020-01-01"


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────

log_records = []

def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)]
    )

def warn(msg, row_data=None):
    logging.warning(msg)
    log_records.append({
        "Nivå": "VARNING",
        "Meddelande": msg,
        "Raddata": str(row_data) if row_data else ""
    })

def info(msg):
    logging.info(msg)
    log_records.append({
        "Nivå": "INFO",
        "Meddelande": msg,
        "Raddata": ""
    })


# ─────────────────────────────────────────────────────────────────────────────
# HJÄLPFUNKTIONER
# ─────────────────────────────────────────────────────────────────────────────

def safe_str(val):
    """Returnera sträng eller tom sträng, aldrig None/nan."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()

def safe_float(val):
    """Returnera float eller None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def parse_date(val):
    """
    Konverterar BSO-datum till ISO 8601 (yyyy-mm-dd).
    BSO kan leverera: Excel-seriellt nummer, datetime-objekt, eller sträng.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        # Excel-seriedatum: dag 1 = 1900-01-01 (Excels felaktiga skottår inräknat)
        try:
            return (date(1899, 12, 30) + __import__("datetime").timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    warn(f"Kunde inte tolka datum: {repr(val)}")
    return None

def map_value(value, mapping, label="värde", default=None):
    """Slå upp värde i en mappning. Logga varning vid miss."""
    key = safe_str(value).upper().strip()
    if key in mapping:
        return mapping[key]
    # Prova originalsträngen (case-sensitive)
    if safe_str(value) in mapping:
        return mapping[safe_str(value)]
    if default is not None:
        warn(f"Okänt {label}: '{value}' — använder fallback '{default}'", {"värde": value})
        return default
    warn(f"Okänt {label}: '{value}' — ingen fallback, raden kan behöva manuell kontroll", {"värde": value})
    return None

def find_sheet(xl: pd.ExcelFile, candidates: list[str]):
    """Hitta rätt flik med fallback-namn."""
    sheets_lower = {s.lower(): s for s in xl.sheet_names}
    for c in candidates:
        if c.lower() in sheets_lower:
            return xl.parse(sheets_lower[c.lower()])
    return None

def save_nuvo_xlsx(df: pd.DataFrame, path: Path, sheet_name: str = "Import"):
    """Sparar en DataFrame som Nuvo-redo xlsx med formaterad rubrikrad."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Autobredd (max 50 tecken)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)
    info(f"Sparad: {path.name}  ({len(df)} rader)")


# ─────────────────────────────────────────────────────────────────────────────
# TRANSFORMATIONER PER ENTITET
# ─────────────────────────────────────────────────────────────────────────────

def transform_timpriser(df: pd.DataFrame) -> pd.DataFrame:
    """
    BSO: SELECT cReskod AS Kod, cBenamning AS Beskrivning, cEnhet AS Enhet,
              nPris AS Kost_per_tim, nPrisUt AS APris, cKonto AS Kontonr
         FROM Resurser WHERE cTyp = 'EGENTID'

    → Next Nuvo: next_artikel (kategori Tidspris)
    Obligatoriska fält: Kod*, Beskrivning*, Enhet*, Kostn/enh*, Pris/enh*, Kontonr*, Debiterbar*
    """
    rows = []
    for _, r in df.iterrows():
        kod = safe_str(r.get("Kod") or r.get("cReskod") or r.get("KOD", ""))
        if not kod:
            warn("Timpriser: rad utan Kod hoppas över", dict(r))
            continue

        rows.append({
            "Kod":         kod,
            "Beskrivning": safe_str(r.get("Beskrivning") or r.get("cBenamning", "")),
            "Enhet":       safe_str(r.get("Enhet") or r.get("cEnhet", "tim")) or "tim",
            "Kostn/enh":   safe_float(r.get("Kost_per_tim") or r.get("nPris")),
            "Pris/enh":    safe_float(r.get("APris") or r.get("nPrisUt")),
            "Kontonr":     safe_str(r.get("Kontonr") or r.get("cKonto", "7010")),
            "Debiterbar":  "Ja",
        })
    result = pd.DataFrame(rows)
    info(f"Timpriser: {len(result)} rader transformerade")
    return result


def transform_prislista(df: pd.DataFrame) -> pd.DataFrame:
    """
    BSO: SELECT cReskod AS Kod, cBenamning AS Beskrivning, cEnhet AS Enhet,
              nPris AS Kost_per_tim, nPrisUt AS APris, cKonto AS Kontonr, cTyp AS Resurstyp
         FROM Resurser WHERE cTyp <> 'EGENTID' OR cTyp IS NULL

    → Next Nuvo: next_artikel (prislista)
    Obligatoriska fält: Artikelnr*, Beskrivning*, Enhet*, Kostn/enh*, Pris/enh*, Konto*
    """
    rows = []
    for _, r in df.iterrows():
        kod = safe_str(r.get("Kod") or r.get("cReskod", ""))
        if not kod:
            warn("Prislista: rad utan Kod hoppas över", dict(r))
            continue

        resurstyp = safe_str(r.get("Resurstyp") or r.get("cTyp", ""))
        kategori = RESURSTYP_MAP.get(resurstyp.upper() if resurstyp else None,
                   RESURSTYP_MAP.get(None))

        rows.append({
            "Artikelnr":       kod,
            "Beskrivning":     safe_str(r.get("Beskrivning") or r.get("cBenamning", "")),
            "Enhet":           safe_str(r.get("Enhet") or r.get("cEnhet", "st")) or "st",
            "Kostn/enh":       safe_float(r.get("Kost_per_tim") or r.get("nPris")),
            "Pris/enh":        safe_float(r.get("APris") or r.get("nPrisUt")),
            "Konto":           safe_str(r.get("Kontonr") or r.get("cKonto", "4010")),
            "Artikelkategori": kategori,
            "Debiterbar":      "Ja",
        })
    result = pd.DataFrame(rows)
    info(f"Prislista: {len(result)} rader transformerade")
    return result


def transform_projekt(df: pd.DataFrame, cutoff_date: str = CUTOFF_DATE_DEFAULT) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    BSO: SELECT 1 AS Vald, Objektindex, Objektnr AS Projektnr, Objekt AS Projektnamn,
              cStatus AS BSStatus, '50' AS Status, '' AS Projektledare,
              '' AS Projektledarenamn, dPrjSta AS Projektstart, cDebForm AS Ersättningsform
         FROM Objekt ... ORDER BY Objektnr

    → Next Nuvo: next_projekt
    Obligatoriska: Föräldraprojektnr*, Projektnr*, Projektnamn*, Projektstatus*,
                   Projektledare*, Projektstart*, Ersättningsform*, Kund*

    Returnerar: (aktiva_projekt_df, hoppade_över_df)

    OBS: Projektledare och Kund är TOMMA i BSO-exporten — konsulten måste fylla
         dessa MANUELLT eller via ett kompletterande kundregister-Excel.
         Scriptet lämnar fälten tomma och varnar.
    """
    aktiva = []
    historiska = []
    cutoff = datetime.strptime(cutoff_date, "%Y-%m-%d").date()

    for _, r in df.iterrows():
        # Vald-flagga: om BSO-exporten har en kryssruta "Vald" och raden är avmarkerad
        vald = r.get("Vald", 1)
        if str(vald).strip() in ("0", "False", "false", "Nej"):
            historiska.append({**dict(r), "_orsak": "Ej vald i BSO-exporten"})
            continue

        projektnr = safe_str(r.get("Projektnr") or r.get("Objektnr", ""))
        if not projektnr:
            warn("Projekt: rad utan Projektnr hoppas över", dict(r))
            continue

        projektstart_raw = r.get("Projektstart") or r.get("dPrjSta")
        projektstart = parse_date(projektstart_raw)

        # Cut-off: projekt utan startdatum eller äldre än cut-off → historisk
        if projektstart:
            start_date = datetime.strptime(projektstart, "%Y-%m-%d").date()
            if start_date < cutoff:
                historiska.append({**dict(r), "_orsak": f"Startdatum {projektstart} är före cut-off {cutoff_date}"})
                continue

        # BSO Status → Next Status
        bso_status = safe_str(r.get("BSStatus") or r.get("cStatus", ""))
        next_status = map_value(bso_status, BSO_STATUS_MAP, "BSO-projektstatus", DEFAULT_PROJECT_STATUS)

        # Ersättningsform
        ers_raw = safe_str(r.get("Ersättningsform") or r.get("cDebForm") or r.get("Ers", ""))
        ersattningsform = map_value(ers_raw, ERSATTNINGSFORM_MAP, "Ersättningsform", "Löpande räkning")

        # Projektledare: ALLTID tom från BSO — varning
        projektledare = safe_str(r.get("Projektledare", ""))
        if not projektledare:
            warn(f"Projekt {projektnr}: Projektledare saknas — MÅSTE fyllas manuellt", {"Projektnr": projektnr})

        # Kund: saknas i standard BSO-export — varning
        kund = safe_str(r.get("Kund") or r.get("Kundnr") or r.get("Kund nr", ""))
        if not kund:
            warn(f"Projekt {projektnr}: Kundnr saknas — MÅSTE fyllas manuellt", {"Projektnr": projektnr})

        aktiva.append({
            "Föräldraprojektnr": "00",       # Kan anpassas om kunden har projektträd i BSO
            "Projektnr":         projektnr,
            "Projektnamn":       safe_str(r.get("Projektnamn") or r.get("Objekt", "")),
            "Projektstatus":     next_status,
            "Projektledare":     projektledare,  # FYLL I: inloggningsnamn i Next
            "Projektstart":      projektstart or "",
            "Projektslut":       parse_date(r.get("Projektslut") or r.get("dPrjSlut")) or "",
            "Ersättningsform":   ersattningsform,
            "Kund":              kund,           # FYLL I: kundnr i Next
            "Noteringar":        f"BSO-status: {bso_status}",
        })

    aktiva_df = pd.DataFrame(aktiva) if aktiva else pd.DataFrame()
    historiska_df = pd.DataFrame(historiska) if historiska else pd.DataFrame()

    info(f"Projekt: {len(aktiva_df)} aktiva, {len(historiska_df)} historiska/överhoppade")
    if len(aktiva_df) > 0:
        tomma_pl = (aktiva_df["Projektledare"] == "").sum()
        tomma_kund = (aktiva_df["Kund"] == "").sum()
        if tomma_pl > 0:
            warn(f"Projekt: {tomma_pl} rader saknar Projektledare — fyll i manuellt")
        if tomma_kund > 0:
            warn(f"Projekt: {tomma_kund} rader saknar Kund — fyll i manuellt")

    return aktiva_df, historiska_df


def transform_ata(df: pd.DataFrame) -> pd.DataFrame:
    """
    BSO: SELECT PAVV.nSerieNr AS avvnr, PAVV.cDocType AS avvdokumenttyp,
              PAVV.cRubrik AS avvrubrik, PAVV.dDatum1 AS avvupprattad,
              PAVV.dDatum2 AS avvreviderad, PAVV.cTxt2 AS avvdebiteringsform,
              SUBQAR.Belopp, PAVV.cPath [= projektnr]...
         FROM PAvvikelser PAVV ...

    → Next Nuvo: next_additionalworkorder
    Obligatoriska: Projektnr*, Beskrivning*, Typ*, Status*, Kostnadsreglering*, Kund*

    OBS: Kostnadsdetaljer (ÄTA-rader) FINNS i BSO:s Avvikelsefält men mappar ej direkt
         till Next ÄTA-orderrader. Beloppet importeras som Kalkylerad summa.
         Produktgap C1 gäller fortfarande för historiska bokförda kostnader per ÄTA.
    """
    rows = []
    for _, r in df.iterrows():
        projektnr = safe_str(r.get("cPath") or r.get("Projektnr", ""))
        rubrik = safe_str(r.get("avvrubrik") or r.get("cRubrik") or r.get("Rubrik", "Ej angiven rubrik"))

        if not projektnr:
            warn(f"ÄTA '{rubrik}': saknar Projektnr — hoppas över", dict(r))
            continue

        # ÄTA-typ
        doc_type = safe_str(r.get("avvdokumenttyp") or r.get("cDocType", "AVVE")).upper()
        ata_typ = AVVTYP_MAP.get(doc_type, "ÄTA")

        # Debiteringsform → Ersättningsform
        deb_raw = safe_str(r.get("avvdebiteringsform") or r.get("cTxt2", ""))
        ersattningsform = map_value(deb_raw, ERSATTNINGSFORM_MAP, "ÄTA-ersättningsform", "Fastpris")

        # Belopp (kalkylerat)
        belopp = safe_float(r.get("Belopp") or r.get("avvkostnad_krr") or r.get("nSingle1"))

        # Datum
        upprattad = parse_date(r.get("avvupprattad") or r.get("dDatum1"))
        reviderad = parse_date(r.get("avvreviderad") or r.get("dDatum2"))

        rows.append({
            "Projektnr":        projektnr,
            "Beskrivning":      rubrik,
            "Typ":              ata_typ,
            "Status":           "Godkänd",  # Importerade ÄTA antas godkända; justera vid behov
            "Ersättningsform":  ersattningsform,
            "Kostnadsreglering": "Mot UE/Beställare",
            "Kund":             "",         # FYLL I: samma kund som projektet om ej annan
            "Upprättad":        upprattad or "",
            "Reviderad":        reviderad or "",
            "Kalkylerat belopp": belopp or "",
            "BSO-typ":          doc_type,   # Behålls som referens
        })

    result = pd.DataFrame(rows)
    info(f"ÄTA: {len(result)} rader transformerade")
    if len(result) > 0:
        tomma_kund = (result["Kund"] == "").sum()
        if tomma_kund:
            warn(f"ÄTA: {tomma_kund} rader saknar Kund — fyll i manuellt")
    return result


def transform_kostnader(df: pd.DataFrame) -> pd.DataFrame:
    """
    BSO: SELECT cVerifNr, dDatum, nVBelopp, cKonto FROM PAvstamrader WHERE nProjID = ...
         (och fullständigare: SELECT nProjID, objekt.objektnr AS Projektnr,
              cPath, dDatum AS Datum, cLevPersNr AS Levid, cFaktNr AS Fakturanummer,
              cMark AS Fakturamark, cKonto, nVBelopp/nSumma...)

    → Next Nuvo: next_kostnad (import 13)
    Kolumner: Projektnr*, Verifikation nr*, Verifikationstext*, Konto*, Bokf datum*, Nettobelopp*

    Filtrerar bort: balansräkningskonton (1xxx, 2xxx, 8xxx, 9xxx)
    """
    rows = []
    skipped = 0

    for _, r in df.iterrows():
        konto = safe_str(r.get("cKonto") or r.get("Konto", ""))
        if not konto:
            skipped += 1
            continue

        # Hoppa över balansräkningskonton
        if konto.lstrip("0")[:1] in SKIP_ACCOUNTS_PREFIX:
            skipped += 1
            continue

        projektnr = safe_str(r.get("Projektnr") or r.get("objektnr", ""))
        belopp = safe_float(r.get("nVBelopp") or r.get("nSumma") or r.get("Belopp"))
        if belopp is None or belopp == 0:
            skipped += 1
            continue

        verif_nr = safe_str(r.get("cVerifNr") or r.get("Fakturanummer", "IMPORT"))
        datum = parse_date(r.get("dDatum") or r.get("Datum"))

        rows.append({
            "Projektnr":           projektnr,
            "AO/ÄTA nr":           "",  # Lämna tomt — BSO har ej per-ÄTA-kostnader (C1-gap)
            "Verifikation nr":     verif_nr,
            "Verifikationstext":   safe_str(r.get("cMark") or r.get("Fakturamark") or f"BSO import {verif_nr}"),
            "Konto":               konto,
            "Bokf datum":          datum or "",
            "Nettobelopp":         belopp,
            "Leverantör":          safe_str(r.get("cLevnr") or r.get("Levid", "")),
            "Leverantörens fakturanr": safe_str(r.get("cFaktNr") or r.get("Fakturanummer", "")),
        })

    result = pd.DataFrame(rows)
    info(f"Kostnader: {len(result)} rader transformerade, {skipped} hoppade över (balansräkning/tomt)")
    return result


def transform_timmar(df: pd.DataFrame) -> pd.DataFrame:
    """
    BSO: SELECT avst.nProjID, o.objektnr AS Projektnr, o.objekt AS Projektnamn,
              avst.nAvvNr AS AO_AWO, avst.cLevPersNr AS Persid, '' AS Persnamn,
              avst.cYrkesKat AS Roll, avst.cArtNrK AS Levnr,
              SUM(avst.nVKvantitet) AS Timmar, SUM(avst.nVBelopp) AS Nettobelopp
         FROM PAvstamrader avst INNER JOIN Objekt o ON ...
         WHERE cTyp IN ('EGENTID') AND cKat LIKE 'V%'

    → Next Nuvo: next_kostnad (konto 7010, enhet tim)
    Tidkostnader importeras som verifikationsrader med tim-konto.

    OBS: Kräver att personen HAR ett Next-konto. Deaktiverade användare (C3-gap)
         importeras mot ett "ARVSRESURS"-konto — konsulten korrigerar manuellt.
    """
    rows = []
    skipped = 0

    for _, r in df.iterrows():
        projektnr = safe_str(r.get("Projektnr") or r.get("objektnr", ""))
        if not projektnr:
            skipped += 1
            continue

        timmar = safe_float(r.get("Timmar") or r.get("nVKvantitet"))
        belopp = safe_float(r.get("Nettobelopp") or r.get("nVBelopp"))
        if timmar is None or timmar == 0:
            skipped += 1
            continue

        persid = safe_str(r.get("Persid") or r.get("cLevPersNr", "OKÄND"))
        roll = safe_str(r.get("Roll") or r.get("cYrkesKat", ""))
        datum = parse_date(r.get("dDatum") or r.get("Datum"))

        rows.append({
            "Projektnr":       projektnr,
            "AO/ÄTA nr":       safe_str(r.get("AO_AWO") or r.get("nAvvNr", "")),
            "Verifikation nr":  f"TIM-{persid}-{datum or 'OKÄNT'}",
            "Verifikationstext": f"Antal timmar: {timmar}  /  {roll or persid}",
            "Konto":            "7010",  # Löner yrkesarbetare; ändra till 7210 för tjänstemän
            "Bokf datum":       datum or "",
            "Nettobelopp":      belopp or 0,
            "Notering":         f"Resurs-ID i BSO: {persid}  /  Roll: {roll}",
        })

    result = pd.DataFrame(rows)
    info(f"Timmar: {len(result)} rader transformerade, {skipped} hoppade över")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# HUVUDFUNKTION
# ─────────────────────────────────────────────────────────────────────────────

def transform(input_path: str, output_dir: str, cutoff_date: str):
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        print(f"Filen hittades inte: {input_path}")
        sys.exit(1)

    info(f"Läser BSO-export: {input_path.name}")
    xl = pd.ExcelFile(input_path)
    info(f"Tillgängliga flikar: {xl.sheet_names}")

    results = {}

    # ── 1. Timpriser ──────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Timpriser", "Tim priser", "Resurser EGENTID", "EGENTID", "Timmar priser"])
    if df is not None:
        results["01_Timpriser"] = transform_timpriser(df)
    else:
        warn("Ingen Timpriser-flik hittades i BSO-exporten. Hoppar över.")

    # ── 2. Prislista ──────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Prislista", "Artiklar", "Resurser", "Resurser Prislista", "A-Prislista"])
    if df is not None:
        results["02_Prislista"] = transform_prislista(df)
    else:
        warn("Ingen Prislista-flik hittades. Hoppar över.")

    # ── 3. Projekt ────────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Projekt", "Grunddata", "Projektlista", "Projects"])
    if df is not None:
        aktiva, historiska = transform_projekt(df, cutoff_date)
        results["03_Projekt"] = aktiva
        if not historiska.empty:
            results["03_Projekt_HISTORISKA_hoppade"] = historiska
    else:
        warn("Ingen Projekt-flik hittades. Hoppar över.")

    # ── 4. ÄTA ────────────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Avvikelser", "ÄTA", "ATA", "Avvikelse", "ÄTA Avvikelse"])
    if df is not None:
        results["04_ÄTA"] = transform_ata(df)
    else:
        warn("Ingen Avvikelser/ÄTA-flik hittades. Hoppar över.")

    # ── 5. Kostnader ──────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Kostnader", "Bokförda kostnader", "PAvstamrader", "Verifikat"])
    if df is not None:
        results["05_Kostnader"] = transform_kostnader(df)
    else:
        warn("Ingen Kostnader-flik hittades. Hoppar över.")

    # ── 6. Timmar ─────────────────────────────────────────────────────────────
    df = find_sheet(xl, ["Timmar", "Bokförda timmar", "Next timmar", "Tider"])
    if df is not None:
        results["06_Timmar"] = transform_timmar(df)
    else:
        warn("Ingen Timmar-flik hittades. Hoppar över.")

    # ── Spara output ──────────────────────────────────────────────────────────
    for filename, df in results.items():
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            warn(f"Ingen data för {filename} — fil skapas ej")
            continue
        out_path = output_dir / f"{filename}.xlsx"
        save_nuvo_xlsx(df, out_path, sheet_name=filename.split("_", 1)[-1][:25])

    # ── Transformationslogg ───────────────────────────────────────────────────
    if log_records:
        log_df = pd.DataFrame(log_records)
        log_path = output_dir / "TRANSFORMATIONSLOGG.xlsx"
        save_nuvo_xlsx(log_df, log_path, sheet_name="Logg")

    # ── Sammanfattning ────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("TRANSFORMERING KLAR")
    print("="*60)
    print(f"Outputkatalog: {output_dir}")
    for filename, df in results.items():
        if isinstance(df, pd.DataFrame) and not df.empty:
            print(f"  ✓ {filename}.xlsx  ({len(df)} rader)")
    print(f"\nTransformationslogg: TRANSFORMATIONSLOGG.xlsx")
    varningar = [r for r in log_records if r["Nivå"] == "VARNING"]
    print(f"Totalt {len(varningar)} varningar — granska loggen!")

    if "03_Projekt" in results and not results["03_Projekt"].empty:
        proj = results["03_Projekt"]
        print("\n" + "─"*40)
        print("MANUELLA STEG KRÄVS:")
        print("─"*40)
        tomma_pl = (proj["Projektledare"] == "").sum()
        tomma_kund = (proj["Kund"] == "").sum()
        if tomma_pl:
            print(f"  ⚠  03_Projekt.xlsx: {tomma_pl} rader saknar Projektledare (inloggningsnamn i Next)")
        if tomma_kund:
            print(f"  ⚠  03_Projekt.xlsx: {tomma_kund} rader saknar Kund (kundnr i Next)")
        print("  ⚠  04_ÄTA.xlsx: Fyll i Kund-kolumnen (BSO exporterar ej kundnr per ÄTA)")
        print("  ⚠  Produktgap C1: ÄTA-detaljkostnader importeras EJ (BSO saknar koppling)")
        print("  ⚠  Deaktiverade användare i Timmar: kontrollera Notering-kolumnen")


if __name__ == "__main__":
    setup_logging()
    parser = argparse.ArgumentParser(
        description="Transformerar BSO Excel-export till Nuvo-färdiga importfiler för Next Project."
    )
    parser.add_argument("input",
                        help="Sökväg till BSO-exportfilen (.xlsx)")
    parser.add_argument("--output-dir", "-o", default="next_import",
                        help="Katalog för output-filer (default: ./next_import)")
    parser.add_argument("--cutoff", "-c", default=CUTOFF_DATE_DEFAULT,
                        help=f"Cut-off-datum för historiska projekt (default: {CUTOFF_DATE_DEFAULT})")
    args = parser.parse_args()

    transform(args.input, args.output_dir, args.cutoff)
